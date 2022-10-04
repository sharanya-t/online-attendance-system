from tkinter import *
from PIL import ImageTk
from tkinter import ttk
from ttkthemes import themed_tk as tk
import tkinter.messagebox
import os
from os import path
from datetime import date
import openpyxl
from openpyxl import load_workbook
import pandas as pd

def sheetnames():
    xl = pd.ExcelFile('attendance/attend.xlsx')
    ls=xl.sheet_names
    return ls

def subname():
    global subject
    subject=e1.get()
    f = open('notepad.txt','w')
    f.truncate(0)
    f.write(subject)
    f.close()
    return subject
    


def home():
    sub= subname()
    lm=sheetnames()
    flag=0
    if(sub==""):
        tkinter.messagebox.showerror('ERROR','ENTER THE SUBJECT')
    else:
        for i in lm:
            if(i==sub):
                flag=1
                break
        if(flag==1):
            
            root.destroy()
            os.system("python main.py")
        else:
            tkinter.messagebox.showerror('ERROR','SUBJECT DOES NOT EXIST')
        
    
def addsub():
    sub=subname()
    ln=sheetnames()
    f=0
    dt=date.today()
    td=dt.strftime('%d.%m.%y')
    fname='attendance/attend_'+td+'.xlsx'
    if(sub==""):
        tkinter.messagebox.showerror('ERROR','ENTER THE SUBJECT')
    else:
        for i in ln:
            if(i==sub):
                f=1
                break
        if(f==1):
            tkinter.messagebox.showerror('ERROR','SUBJECT ALREADY EXISTS')
        else:
            if path.exists(fname):
                wb2=openpyxl.load_workbook(fname)
                source = wb2['gg']
                wb2.copy_worksheet(source)
                wb2_sheet = wb2['gg Copy']
                wb2_sheet.title = sub
                wb2.save(fname)
            wb=openpyxl.load_workbook('attendance/attend.xlsx')
            source = wb['gg']
            wb.copy_worksheet(source)
            wb_sheet = wb['gg Copy']
            wb_sheet.title = sub
            wb.save('attendance/attend.xlsx')
            tkinter.messagebox.showinfo('SUCCESS','SUBJECT ADDED')

root = tk.ThemedTk()
root.geometry("640x480")
root.title("Online Attendance") 
root.iconbitmap(r'att.ico')

root.get_themes()
root.set_theme("breeze")

statusbar=ttk.Label(root, text="Login Page", relief=SUNKEN, anchor=W)
statusbar.pack(side=BOTTOM, fill=X)

topframe = Frame(root)
topframe.pack(padx=20,pady=20)

mvitPhoto = ImageTk.PhotoImage(file='mvitlogo.png')
mvitico=ttk.Label(topframe, image = mvitPhoto)
mvitico.grid(row=0,column=1,padx=10)

filelabel = ttk.Label(topframe, text = 'LOGIN PAGE', font="Times 20 underline")
filelabel.grid(row=1,column=1, pady=20)

middleframe = Frame(root)
middleframe.pack(pady=10)

sublabel = ttk.Label(middleframe, text = 'SUBJECT', font="Times 15 normal")
sublabel.pack(side=LEFT,padx=10)

e1=ttk.Entry(middleframe)
e1.pack(padx=10)
subname()

bottomframe = Frame(root)
bottomframe.pack(pady=20)

loginphoto = PhotoImage(file='login.png')
loginBtn = ttk.Button(bottomframe, image = loginphoto, command=home)
loginBtn.grid(row=0,column=0,padx=20,pady=20)

loginlabel = ttk.Label(bottomframe, text = 'LOGIN', font="Times 11 normal")
loginlabel.grid(row=1,column=0,padx=20)

addphoto = PhotoImage(file='addsub.png')
addBtn = ttk.Button(bottomframe,image = addphoto,command=addsub)
addBtn.grid(row=0,column=1,padx=20,pady=20)

addlabel = ttk.Label(bottomframe, text = 'ADD SUBJECT', font="Times 11 normal")
addlabel.grid(row=1,column=1,padx=20)

root.mainloop()
