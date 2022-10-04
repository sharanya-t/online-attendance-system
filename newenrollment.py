from tkinter import *
import os
import cv2
from time import sleep
import openpyxl
from openpyxl import load_workbook
import csv
from datetime import date
from os import path
from tkinter import ttk
import tkinter.messagebox
from ttkthemes import themed_tk as tk
import pandas as pd

def sheetnames():
    xl = pd.ExcelFile('attendance/attend.xlsx')
    ls=xl.sheet_names
    return ls

def home():
    root.destroy()
    os.system('python main.py')


def appendtoexcel():
    lm=sheetnames()
    dt=date.today()
    td=dt.strftime('%d.%m.%y')
    fname='attendance/attend_'+td+'.xlsx'
    if path.exists(fname):
        wb2=openpyxl.load_workbook(fname)
        for i in lm:
            sheet=wb2[i]
            usn=e1.get()
            usn = str(usn)
            name=e2.get()
            lastthree=int(usn[-3:])
            for i in range(2,sheet.max_row+1):
                val = sheet.cell(row=i,column=1).value
                val = str(val)
                ltval = int(val[-3:])
                if ltval>lastthree:
                    break
            sheet.insert_rows(i,1)
            sheet.cell(row=i,column=1).value=usn
            sheet.cell(row=i,column=2).value=name
            sheet.cell(row=i,column=3).value='absent'    
        wb2.save(fname)
        wb2.close()
    wb=openpyxl.load_workbook("attendance/attend.xlsx")
    for i in lm:
        sheet=wb[i]
        usn=e1.get()
        usn = str(usn)
        name=e2.get()
        lastthree=int(usn[-3:])
        for i in range(2,sheet.max_row+1):
           val = sheet.cell(row=i,column=1).value
           val = str(val)
           ltval = int(val[-3:])
           if ltval>lastthree:
               break
        sheet.insert_rows(i,1)
        sheet.cell(row=i,column=1).value=usn
        sheet.cell(row=i,column=2).value=name
        sheet.cell(row=i,column=3).value='absent'
    wb.save("attendance/attend.xlsx")
    wb.close()
    tkinter.messagebox.showinfo('Online Attendance','REGISTERED SUCCESSFULLY! :)')
    return 

def save_pic():
        webcam = cv2.VideoCapture(0)
        sleep(2)
        name=e1.get()
        str=name+'.jpg'
        path='people'
        while True:
            check, frame = webcam.read()   
            cv2.imshow("Capturing", frame)
            key = cv2.waitKey(1)
            
            if key == ord('s'): 
                cv2.imwrite(os.path.join(path,str), img=frame)
                
                print("Image saved!")
                webcam.release()
                cv2.destroyAllWindows()
                tkinter.messagebox.showinfo('Online Attendance','Picture Saved!')
                break

            elif key == ord('q'):
                webcam.release()
                cv2.destroyAllWindows()
                tkinter.messagebox.showerror('Online Attendance','No picture clicked!')
                break

root = tk.ThemedTk()
root.geometry("640x480")
root.title("Enrollment Window") 
root.iconbitmap(r'att.ico')

root.get_themes()
root.set_theme("breeze")

menubar = Menu(root)
root.config(menu=menubar)
subMenu = Menu(menubar,tearoff=0)
menubar.add_command(label="HOME",command=home)

statusbar=ttk.Label(root, text="Enrolling..", relief=SUNKEN, anchor=W)
statusbar.pack(side=BOTTOM, fill=X)

topframe = Frame(root)
topframe.pack(padx=20,pady=20)

filelabel = ttk.Label(topframe, text = 'ENROLLMENT', font="Times 20 underline")
filelabel.pack(pady=10)

deets = ttk.Label(topframe, text = 'Student Details', font="Times 15 normal")
deets.pack()

middleframe = Frame(root)
middleframe.pack(pady=10)

enrolllabel = ttk.Label(middleframe, text = 'USN', font="Times 15 normal")
enrolllabel.pack(side=LEFT, padx=20)

e1=ttk.Entry(middleframe)
e1.pack()

nextframe = Frame(root)
nextframe.pack(pady=10)

usnlabel = ttk.Label(nextframe, text = 'NAME', font="Times 15 normal")
usnlabel.pack(side=LEFT, padx=20)

e2=ttk.Entry(nextframe)
e2.pack()

bottomframe = Frame(root)
bottomframe.pack()

marklabel = ttk.Label(bottomframe, text = 'TAKE PHOTO', font="Times 15 normal")
marklabel.pack(side=LEFT, padx=20)

cam = PhotoImage(file='camera.png')
camBtn = ttk.Button(bottomframe, image = cam, command=save_pic)
camBtn.pack(pady=30)

lastframe = Frame(root)
lastframe.pack()

btn1 = ttk.Button(lastframe, text='Register',command=appendtoexcel)
btn1.pack()

root.mainloop()

