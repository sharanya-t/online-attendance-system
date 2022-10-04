import openpyxl
from datetime import date
import os.path
from os import path

def mark_attendance(name):

    dt=date.today()
    td=dt.strftime('%d.%m.%y')
    fname='attendance/attend_'+td+'.xlsx'
    

    wb=openpyxl.load_workbook('attendance/attend.xlsx')
    sheet=wb.active
    

    for i in range(1,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value==name):
            break
    
    # print(i)
    if(i<sheet.max_row+1):
        sheet.cell(row=i,column=2).value='present'
    
    

    wb.save(filename=fname)
    wb.close()

def mark_sub_att(name,sub):
    dt=date.today()
    td=dt.strftime('%d.%m.%y')
    fname='attendance/attend_'+td+'.xlsx'
    if path.exists(fname):
        wb=openpyxl.load_workbook(fname)
    else:
        wb=openpyxl.load_workbook('attendance/attend.xlsx')
    
    
    sheet=wb[sub]

    for i in range(2,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value==name):
            break
    
    # print(i)
    if(i<sheet.max_row+1):
        sheet.cell(row=i,column=3).value='present'

    wb.save(filename=fname)
    wb.close()


